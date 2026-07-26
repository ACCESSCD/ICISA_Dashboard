#!/usr/bin/env python3
"""
Run this script whenever the Faculty list Excel is updated.
Writes speakers.json (excluding Declined speakers) for the dashboard.

Usage:  python generate_data.py
"""
import json
import re
import unicodedata
import warnings
from collections import Counter
from pathlib import Path

import openpyxl

_ICISA_DIR   = Path(r'C:\Users\carol\PycharmProjects\ICISA information')
_candidates  = sorted(_ICISA_DIR.glob('Faculty list follow up*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
if not _candidates:
    raise FileNotFoundError(f'No Faculty list follow up*.xlsx found in {_ICISA_DIR}')
FACULTY_PATH = _candidates[0]
print(f'Using faculty file: {FACULTY_PATH.name}')
_pgm_candidates = sorted(_ICISA_DIR.glob('PGM ICISA*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
if not _pgm_candidates:
    raise FileNotFoundError(f'No PGM ICISA*.xlsx found in {_ICISA_DIR}')
PGM_PATH     = _pgm_candidates[0]
print(f'Using programme file: {PGM_PATH.name}')
OUTPUT_PATH  = Path(__file__).parent / 'speakers.json'

EXCLUDE_TYPES = {'DO NOT INVITE'}

# Speakers removed at organiser request (not yet marked Declined in Excel)
MANUAL_EXCLUDE = {
    ('meg',   'rosenblatt'),
    ('vaida', 'sonia'),
    ('ron',   'george'),
    ('steffen', 'rex'),
}

# Known programme discrepancies that the automatic name-matching can't catch
# (first name too short for fuzzy detection, or the programme spelling isn't a
# near-miss of the faculty-list spelling). Correct spelling is always the
# Faculty list Excel; this dict just documents the mismatch so it shows up in
# the dashboard's "Spelling issues" panel instead of "no sessions".
MANUAL_SPELLING_FLAGS = {
    ('jacob', 'refael'):
        'Programme lists as "Jacob Raphael" (Blood session – POC-based '
        'transfusion algorithms talk). Faculty list spelling "Refael" is correct.',
    ('david', 'polaner'):
        'Has a talk in the Pediatrics session but no title is listed in the '
        'programme (row shows "David Polaner" with "?" where the title should be).',
    ('richebe', 'philippe'):
        'First/last name are swapped in the Faculty list Excel (first-name '
        'column has "Richebe", last-name column has "Philippe") — correct '
        'name is Philippe Richebe. He does have 2 genuine sessions (a Pain '
        'talk and a debate slot); only the name order is wrong.',
}

# Minimum first-name length for spelling-mismatch detection.
# Short names (e.g. "David", "Nadav") are too common and cause false positives.
MIN_FIRST_NAME_LEN = 6

# Auto-detected spelling flags that are false positives (reviewed and cleared
# by the organiser) — fuzzy name matching can pick up an unrelated speaker's
# name that happens to be similar (e.g. "Philipp" Lirk vs. "Philippe" Richebe).
FALSE_POSITIVE_SPELLING = {
    ('philipp', 'lirk'),
}


# ── helpers ─────────────────────────────────────────────────────────────────

def clean(v):
    return str(v).strip() if v is not None else ''


def norm(s):
    """Lowercase + strip diacritics for fuzzy matching."""
    s = str(s).lower().strip()
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def match_key(last_name):
    """
    Return the best single token to search for a speaker by last name.
    Handles hyphenated names (Haylock-Loor -> haylock) and
    multi-word last names (Gama de Abreu -> abreu).
    """
    n = norm(last_name)
    parts = [p for p in re.split(r'[-\s]+', n) if len(p) > 3]
    if not parts:
        return n
    return max(parts, key=len)   # longest token = most distinctive


def get_status(row):
    s = norm(clean(row[20]) if len(row) > 20 else '')
    c = norm(clean(row[16]) if len(row) > 16 else '')
    if any(x in s for x in ('confirm', 'will come', 'unofficially', 'unofficialy')):
        return 'Confirmed'
    if any(x in s for x in ('declin', 'probably not', 'probaly', 'probably wont',
                              'probaby', 'wont come')):
        return 'Declined'
    if 'declin' in c:
        return 'Declined'
    return 'Pending'


# ── load speakers ────────────────────────────────────────────────────────────

def _is_red_row(cells):
    """Return True if any of the first 10 cells has a solid red background fill."""
    for cell in cells[:10]:
        fill = cell.fill
        if fill and fill.fill_type == 'solid':
            color = fill.fgColor
            if color.type == 'rgb' and color.rgb not in ('00000000', 'FF000000'):
                r = int(color.rgb[2:4], 16)
                g = int(color.rgb[4:6], 16)
                b = int(color.rgb[6:8], 16)
                if r > 180 and g < 80 and b < 80:
                    return True
    return False


def load_speakers():
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(FACULTY_PATH, data_only=True)

    ws = wb['International Speakers']
    speakers, seen = [], set()

    for row in ws.iter_rows(min_row=3):
        if ws.row_dimensions[row[0].row].hidden:
            continue
        if _is_red_row(row):
            continue
        row = tuple(cell.value for cell in row)
        first = clean(row[5])
        last  = clean(row[6])
        if not first or first == 'None':
            continue
            
        # Strip common titles
        first = re.sub(r'^(Prof\.|Prof\s|Dr\.|Dr\s)', '', first, flags=re.IGNORECASE).strip()
        last = re.sub(r'^(Prof\.|Prof\s|Dr\.|Dr\s)', '', last, flags=re.IGNORECASE).strip()

        if 'moderator' in f'{first} {last}'.lower():
            continue

        stype = clean(row[3])
        if stype.upper() in EXCLUDE_TYPES:
            continue

        status = get_status(row)
        if status == 'Declined':
            continue

        if (norm(first), norm(last)) in MANUAL_EXCLUDE:
            continue

        key = (norm(first), norm(last))
        if key in seen:
            continue
        seen.add(key)

        speakers.append({
            'first':          first,
            'last':           last,
            'type':           stype,
            'track':          clean(row[4]),
            'country':        clean(row[7]),
            'email':          clean(row[9]),
            'affil':          clean(row[13]),
            'inv':            'Yes' if (len(row) > 14 and clean(row[14]).upper() == 'V') else '',
            'status':         status,
            'bio':            'Yes' if (len(row) > 21 and row[21]) else '',
            'photo':          'Yes' if (len(row) > 22 and row[22]) else '',
            'tasks':          0,          # filled below
            'spelling_issue': '',         # filled below
        })

    wb.close()
    return speakers


# ── count tasks + flag spelling issues from programme ────────────────────────

_MOD_PREFIX = re.compile(r'^moderator\s*:?\s*', re.I)
_MOD_ANY    = re.compile(r'moderator', re.I)
_TIME_RE    = re.compile(r'^\d{1,2}:\d{2}')
_CFW_PREFIX = re.compile(r'^CFW\s*:\s*', re.I)
_DASH_RE    = re.compile(r'\s[-–]\s')
# Debate-format lines end in "Label: ... : Speaker Name" (e.g. "Con: Keep AI
# Out of My OR: Jens Meier") — no dash, so _DASH_RE misses them. Match a
# trailing colon followed by a short Title-Case name (2-4 words).
_COLON_NAME_RE = re.compile(r":\s*([A-Z][\w.'’-]*(?:\s+[A-Z][\w.'’-]*){1,3})\s*$")


def _title_key(title):
    """First 5 normalised words of a title; used for cross-cell deduplication."""
    return ' '.join(norm(title).split()[:5])


def _add(short_lines, all_raw, all_norm, text, title_key='', for_counting=True):
    nl = norm(text)
    if for_counting:
        short_lines.append((nl, title_key))
    all_raw.append(text)
    all_norm.append(nl)


def _collect_programme_lines():
    """
    Return:
      short_lines — list of (norm_text, title_key) tuples for task counting.
                    title_key is the first 5 words of the associated talk title,
                    used to deduplicate the same talk appearing in multiple cells.
      all_raw / all_norm — broader parallel lists used for spelling detection.
      long_norm   — normalised lines of 6+ words, for second-pass name matching.
    """
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(PGM_PATH, data_only=True)
    ws = wb['Sheet1']

    short_lines = []
    all_raw, all_norm = [], []
    long_norm = []

    for row in ws.iter_rows():
        if ws.row_dimensions[row[0].row].hidden:
            continue
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            if text in ('None', ''):
                continue

            current_title = ''  # most recent long line (likely a talk title)
            cell_header   = ''  # first visible line of cell (session/category)

            for line in text.split('\n'):
                line = line.strip()
                if not line or _TIME_RE.match(line):
                    continue

                # Strip planning-note prefix ("CFW: ...")
                line = _CFW_PREFIX.sub('', line).strip()
                if not line:
                    continue

                if not cell_header:
                    cell_header = line

                # Moderator lines: extract name after the prefix
                if _MOD_ANY.search(line):
                    name_part = _MOD_PREFIX.sub('', line).strip()
                    if name_part and 1 <= len(name_part.split()) <= 4:
                        tk = _title_key(current_title or cell_header)
                        _add(short_lines, all_raw, all_norm, name_part, tk)
                    continue

                words = line.split()

                # "Title - Speaker Name" or "Speaker - Title": extract each
                # segment. The first segment is used as the talk title key.
                # continue prevents double-counting via comma/short-line paths.
                if _DASH_RE.search(line):
                    segments = _DASH_RE.split(line)
                    seg0   = segments[0].strip()
                    seg_tk = (_title_key(seg0)
                              if len(seg0.split()) >= 2
                              else _title_key(current_title or cell_header))
                    for seg in segments:
                        seg = seg.strip()
                        seg_splits = re.split(r',|\s&\s|\sand\s', seg, flags=re.IGNORECASE)
                        if len(seg_splits) > 1:
                            for part in seg_splits:
                                part = part.strip()
                                if part and 1 <= len(part.split()) <= 4:
                                    _add(short_lines, all_raw, all_norm, part, seg_tk)
                                elif part and len(part.split()) <= 8:
                                    _add(short_lines, all_raw, all_norm, part, seg_tk,
                                         for_counting=False)
                        else:
                            seg_words = seg.split()
                            if 1 <= len(seg_words) <= 4:
                                _add(short_lines, all_raw, all_norm, seg, seg_tk)
                            elif len(seg_words) <= 8:
                                _add(short_lines, all_raw, all_norm, seg, seg_tk,
                                     for_counting=False)
                            else:
                                long_norm.append(norm(seg))
                    continue

                # "Label: Title: Speaker Name" debate-format lines: split off
                # the trailing name so it's matched like any other short line.
                colon_m = _COLON_NAME_RE.search(line) if len(words) > 5 else None
                if colon_m:
                    name_part = colon_m.group(1).strip()
                    prefix = line[:colon_m.start()].strip()
                    seg_tk = _title_key(prefix or current_title or cell_header)
                    _add(short_lines, all_raw, all_norm, name_part, seg_tk)
                    current_title = prefix or current_title
                    continue

                # Comma/ampersand-separated roster of short names (e.g. a panel
                # list like "Ruth Landau, Brian Bateman, ..."): treat as names
                # attached to the CURRENT session, not a new talk title.
                # Otherwise someone named earlier in the same cell (e.g. via
                # "Moderator: X") gets counted twice under two different
                # title keys — one from the moderator line, one from this
                # list being mistaken for a >5-word talk title.
                roster_splits = [p.strip() for p in
                                  re.split(r',|\s&\s|\sand\s', line, flags=re.IGNORECASE)
                                  if p.strip()]
                is_roster = (len(roster_splits) >= 3
                             and all(1 <= len(p.split()) <= 4 for p in roster_splits))
                if is_roster:
                    tk = _title_key(current_title or cell_header)
                    for part in roster_splits:
                        _add(short_lines, all_raw, all_norm, part, tk)
                    continue

                # Lines longer than 5 words are likely talk titles — remember
                # them so they can be used as title_key for the name that follows.
                if len(words) > 5:
                    current_title = line

                tk = _title_key(current_title or cell_header)

                # Comma/ampersand-separated panel lists: split and add each part
                line_splits = re.split(r',|\s&\s|\sand\s', line, flags=re.IGNORECASE)
                if len(line_splits) > 1 and len(words) <= 15:
                    for part in line_splits:
                        part = part.strip()
                        if part and 1 <= len(part.split()) <= 3:
                            _add(short_lines, all_raw, all_norm, part, tk)

                # Normal short lines
                if 1 <= len(words) <= 5:
                    _add(short_lines, all_raw, all_norm, line, tk)
                elif len(words) <= 8:
                    _add(short_lines, all_raw, all_norm, line, tk, for_counting=False)
                else:
                    # Long lines kept separately for start-of-line name matching
                    long_norm.append(norm(line))

    wb.close()
    return short_lines, all_raw, all_norm, long_norm


def count_tasks_and_flag(speakers):
    from difflib import SequenceMatcher

    short_lines, all_raw, all_norm, long_norm = _collect_programme_lines()

    # ── Task counting pass 1: first-name + last-name + title deduplication ─────
    # Always require both first and last name so distinct people with the same
    # last name are never conflated. Deduplicate by title_key so the same talk
    # listed in multiple programme cells counts as one task.
    for s in speakers:
        key = match_key(s['last'])
        if not key:
            continue
        key_re = re.compile(r'\b' + re.escape(key) + r'\b')
        fn = norm(s['first'])
        # Short first names need a word boundary on both sides (e.g. "Ed" must
        # not match "Edmond"). Longer names use a prefix match with no trailing
        # boundary so "Dan" matches "Danny" / "Daniel" interchangeably.
        if len(fn) <= 3:
            fp_re = re.compile(r'\b' + re.escape(fn) + r'\b')
        else:
            fp_re = re.compile(r'\b' + re.escape(fn[:3]))
        seen_titles: set = set()
        for nl, title_key in short_lines:
            if key_re.search(nl) and fp_re.search(nl):
                seen_titles.add(title_key)
        s['tasks'] = len(seen_titles)

    # ── Task counting pass 2: long lines that START with "FirstName LastName…" ─
    # Catches format: "Speaker Name Full Talk Title Here" (name leads the cell).
    for s in speakers:
        if s['tasks'] > 0:
            continue
        fn  = norm(s['first'])
        key = match_key(s['last'])
        if not fn or not key:
            continue
        # Line must begin with the first name and contain the last-name key
        # within the first 5 words (to avoid matching unrelated long text).
        start_re = re.compile(r'^' + re.escape(fn) + r'\b')
        key_re   = re.compile(r'\b' + re.escape(key) + r'\b')
        for nl in long_norm:
            if start_re.match(nl) and key_re.search(' '.join(nl.split()[:5])):
                s['tasks'] += 1
                break

    # ── Spelling mismatch detection ──────────────────────────────────────────
    # Only check speakers with 0 tasks. For each, look for their first name
    # (exact or fuzzy ≥0.85 similarity) in programme lines that lack their
    # correct last-name key. Fuzzy matching catches "Laslo" ≈ "Laszlo".
    for s in speakers:
        if s['tasks'] > 0:
            continue
        fn = norm(s['first'])
        ln = match_key(s['last'])

        if len(fn) < MIN_FIRST_NAME_LEN:
            continue

        fn_re = re.compile(r'\b' + re.escape(fn) + r'\b')
        ln_re = re.compile(r'\b' + re.escape(ln) + r'\b') if ln else None

        for raw, nl in zip(all_raw, all_norm):
            # Exact first-name match
            first_name_hit = fn_re.search(nl)
            # Fuzzy match: any word in the line is ≥0.85 similar to first name
            if not first_name_hit:
                words = re.findall(r'\b\w+\b', nl)
                first_name_hit = any(
                    len(w) >= MIN_FIRST_NAME_LEN - 1
                    and SequenceMatcher(None, fn, w).ratio() >= 0.85
                    for w in words
                )
            if not first_name_hit:
                continue
            if ln_re and ln_re.search(nl):
                continue
            s['spelling_issue'] = raw
            break

    # ── Manual overrides for known mismatches the heuristics above miss ────
    for s in speakers:
        key = (norm(s['first']), norm(s['last']))
        if key in MANUAL_SPELLING_FLAGS and not s['spelling_issue']:
            s['spelling_issue'] = MANUAL_SPELLING_FLAGS[key]

    # ── Clear known false-positive auto-detections ──────────────────────────
    for s in speakers:
        key = (norm(s['first']), norm(s['last']))
        if key in FALSE_POSITIVE_SPELLING:
            s['spelling_issue'] = ''


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    speakers = load_speakers()
    count_tasks_and_flag(speakers)

    total    = len(speakers)
    confirmed = sum(1 for s in speakers if s['status'] == 'Confirmed')
    pending   = sum(1 for s in speakers if s['status'] == 'Pending')
    no_tasks  = sum(1 for s in speakers if s['tasks'] == 0 and not s['spelling_issue'])
    spelling  = sum(1 for s in speakers if s['spelling_issue'])
    heavy     = sum(1 for s in speakers if s['tasks'] >= 4)

    print(f'Speakers (excl. Declined): {total}')
    print(f'  Confirmed      : {confirmed}')
    print(f'  Pending        : {pending}')
    print(f'  No tasks       : {no_tasks}')
    print(f'  Spelling issues: {spelling}')
    print(f'  4+ tasks       : {heavy}')
    print()
    print('Spelling issues:')
    for s in sorted((s for s in speakers if s['spelling_issue']), key=lambda x: x['last']):
        print(f"  {s['first']} {s['last']}  ->  \"{s['spelling_issue']}\"")
    print()
    print('Task counts per speaker:')
    for s in sorted(speakers, key=lambda x: -x['tasks']):
        flag = ' [SPELLING]' if s['spelling_issue'] else ''
        print(f"  {s['tasks']:2d}  {s['first']} {s['last']}  [{s['status']}]{flag}")

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(speakers, f, ensure_ascii=False, indent=2)
    print(f'\nWritten -> {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
