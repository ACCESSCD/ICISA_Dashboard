#!/usr/bin/env python3
"""
Reads CFW follow up.xlsx (sheet "$updates") and writes sponsorship.json
for the dashboard.

Column layout ("$updates" sheet):
  B  — company name
  F  — 2026 comment / follow-up note
  G  — 2026 committed amount (Euro)
  H  — 2026 committed amount (NIS)
  I  — per-company "Minimum Expected" amount, tracked in EURO (per the I1
       header). Row 77 holds a =SUM(...) formula labelled "Expected total",
       but this script sums column I itself (rows 2..that row) rather than
       reading the formula's cached value, since openpyxl edits don't
       trigger Excel to recalculate — reading the cached total would go
       stale after any script-driven edit to the sheet.

Rules (per column, G and H are independent):
  red-flagged → row B:I has a solid red fill -> excluded entirely (company is
                dead/declined), regardless of G/H/I values
  committed   → G and/or H is a positive number
  declined    → G = 0 or H = 0 (excluded entirely)
  skip        → nothing else applies

Besides the committed list, two follow-up lists are derived from comparing
each row's committed amount (G/H, converted to a EUR-equivalent) against its
column-I "Minimum Expected" figure:
  pending_no_commitment → nothing committed yet (G/H both empty/zero) but I
                          has a non-zero figure — a live prospect worth
                          chasing for an answer.
  expected_discrepancy  → something IS committed, but I differs from the
                          committed EUR-equivalent — worth following up to
                          find out whether more is coming (I > committed) or
                          the committed figure needs correcting (I < committed).

If a red-flagged row has a nonzero G, H, or I value, it's printed as a
"red-flagged but non-zero" warning — that combination usually means the
row was highlighted red after money was already logged, so it's worth
the organiser double-checking rather than silently dropping it.

REQUIRED_NIS is the sponsorship target for the conference, set by the
organisers — it isn't tracked anywhere in the spreadsheet, so it's a
constant here. Update it by hand if the target changes.

The summary totals (required / expected / committed / gap) are all
displayed in NIS. Expected Total and Committed are SEPARATE figures,
shown side by side on the dashboard — do not add them together.
Expected Total = the "Expected total" row (column I, in EUR) converted
to NIS. Column I is a per-company "minimum expected" estimate that
already includes most committed sponsors (e.g. Medtronic's committed
€35,000 is also its I value), so summing it with Committed double-counts
those sponsors. Ask the user for the current EUR->NIS rate each time
this is regenerated — it's a hardcoded constant, not read from the
spreadsheet.

Usage:  python generate_sponsorship.py
"""
import json
import sys
import warnings
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

_ICISA_DIR = Path(r'C:\Users\carol\PycharmProjects\ICISA information')
SPONSOR_PATH = _ICISA_DIR / 'CFW follow up.xlsx'
OUTPUT_PATH  = Path(__file__).parent / 'sponsorship.json'
SHEET_NAME   = '$updates'

REQUIRED_NIS = 938423
EUR_TO_NIS   = 3.46

NAME_COL, NOTE_COL, EURO_COL, NIS_COL = 2, 6, 7, 8
EXPECTED_TOTAL_COL = 9  # column I


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _cell_is_red(cell):
    fill = cell.fill
    if fill and fill.fill_type == 'solid':
        color = fill.fgColor
        if color.type == 'rgb' and color.rgb not in ('00000000', 'FF000000'):
            r = int(color.rgb[2:4], 16)
            g = int(color.rgb[4:6], 16)
            b = int(color.rgb[6:8], 16)
            return r > 180 and g < 80 and b < 80
    return False


def _is_red_row(cells):
    """A company is dead only when its NAME cell (column B) is filled solid red.
    Stray red on other columns (e.g. an old note highlight) does not kill the row
    - ACT has red on C/D only but is still a live prospect."""
    return _cell_is_red(cells[NAME_COL - 1])


def load_sponsorship():
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(SPONSOR_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    pipeline_total_eur = 0.0

    committed = []
    pending_no_commitment = []
    expected_discrepancy = []
    red_flagged_nonzero = []

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, NAME_COL).value
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            break
        if name.lower().startswith('expected total'):
            break  # reached the summary row (a =SUM formula) — stop reading sponsor rows

        row_cells = [ws.cell(r, c) for c in range(1, 11)]
        if _is_red_row(row_cells):
            g_raw = _to_float(ws.cell(r, EURO_COL).value) or 0.0
            h_raw = _to_float(ws.cell(r, NIS_COL).value) or 0.0
            i_raw = _to_float(ws.cell(r, EXPECTED_TOTAL_COL).value) or 0.0
            if g_raw or h_raw or i_raw:
                red_flagged_nonzero.append({
                    'name': name, 'row': r, 'euro': g_raw, 'nis': h_raw, 'expected_eur': i_raw,
                })
            continue  # red-flagged rows are excluded entirely from totals

        expected_eur = _to_float(ws.cell(r, EXPECTED_TOTAL_COL).value) or 0.0
        pipeline_total_eur += expected_eur

        note = ws.cell(r, NOTE_COL).value
        note = str(note).strip() if note else ''

        g = _to_float(ws.cell(r, EURO_COL).value)
        h = _to_float(ws.cell(r, NIS_COL).value)

        euro_amt = g if (g is not None and g > 0) else None
        nis_amt  = h if (h is not None and h > 0) else None
        declined = (g == 0) or (h == 0)
        if declined:
            continue

        committed_eur_equiv = (euro_amt or 0) + (nis_amt or 0) / EUR_TO_NIS

        if euro_amt is not None or nis_amt is not None:
            combined_nis = (euro_amt or 0) * EUR_TO_NIS + (nis_amt or 0)
            committed.append({
                'name': name,
                'euro': euro_amt,
                'nis': nis_amt,
                'combined_nis': combined_nis,
            })
            if expected_eur and abs(expected_eur - committed_eur_equiv) > 100:
                expected_discrepancy.append({
                    'name': name,
                    'committed_euro': euro_amt,
                    'committed_nis': nis_amt,
                    'committed_combined_nis': combined_nis,
                    'expected_eur': expected_eur,
                    'expected_nis': expected_eur * EUR_TO_NIS,
                    'delta_nis': expected_eur * EUR_TO_NIS - combined_nis,
                    'note': note,
                })
        elif expected_eur:
            pending_no_commitment.append({
                'name': name,
                'expected_eur': expected_eur,
                'expected_nis': expected_eur * EUR_TO_NIS,
                'note': note,
            })

    wb.close()

    committed.sort(key=lambda x: -x['combined_nis'])
    pending_no_commitment.sort(key=lambda x: -x['expected_eur'])
    expected_discrepancy.sort(key=lambda x: -abs(x['delta_nis']))
    committed_delta_nis = sum(c['combined_nis'] for c in committed)
    expected_total_nis = pipeline_total_eur * EUR_TO_NIS
    gap_nis = REQUIRED_NIS - committed_delta_nis

    return {
        'committed': committed,
        'pending_no_commitment': pending_no_commitment,
        'expected_discrepancy': expected_discrepancy,
        'required_nis': REQUIRED_NIS,
        'expected_total_nis': expected_total_nis,
        'committed_delta_nis': committed_delta_nis,
        'gap_nis': gap_nis,
        'eur_to_nis': EUR_TO_NIS,
        'red_flagged_nonzero': red_flagged_nonzero,
    }


def main():
    data = load_sponsorship()
    red_flagged_nonzero = data.pop('red_flagged_nonzero')

    if red_flagged_nonzero:
        print('WARNING: red-flagged rows with a non-zero value (excluded from totals, please double-check):')
        for w in red_flagged_nonzero:
            parts = []
            if w['euro']:
                parts.append(f"G=€{w['euro']:,.0f}")
            if w['nis']:
                parts.append(f"H=₪{w['nis']:,.0f}")
            if w['expected_eur']:
                parts.append(f"I=€{w['expected_eur']:,.0f}")
            print(f"  row {w['row']:>2}  {w['name']:30} {', '.join(parts)}")
        print()

    print(f'Committed ({len(data["committed"])} companies):')
    for c in data['committed']:
        parts = []
        if c['euro'] is not None:
            parts.append(f'€{c["euro"]:,.0f}')
        if c['nis'] is not None:
            parts.append(f'₪{c["nis"]:,.0f}')
        print(f'  {" + ".join(parts):20} {c["name"]}')
    print(f'  Combined (NIS, @{data["eur_to_nis"]} NIS/EUR): ₪{data["committed_delta_nis"]:,.0f}')
    print()
    print(f'Pending, no commitment yet ({len(data["pending_no_commitment"])} companies):')
    for t in data['pending_no_commitment']:
        print(f'  €{t["expected_eur"]:,.0f}'.ljust(12) + f'{t["name"]:30} {t["note"]}')
    print()
    print(f'Expected/committed discrepancies ({len(data["expected_discrepancy"])} companies):')
    for t in data['expected_discrepancy']:
        print(f'  committed ₪{t["committed_combined_nis"]:,.0f} vs expected ₪{t["expected_nis"]:,.0f}'.ljust(45)
              + f'{t["name"]:30} {t["note"]}')
    print()
    print(f'Required:       ₪{data["required_nis"]:,.0f}')
    print(f'Expected total: ₪{data["expected_total_nis"]:,.0f}')
    print(f'Committed:      ₪{data["committed_delta_nis"]:,.0f}')
    print(f'Remaining gap:  ₪{data["gap_nis"]:,.0f}')

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'\nWritten -> {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
